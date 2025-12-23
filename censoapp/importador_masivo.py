"""
Utilidades para importación masiva de datos desde Excel.
Permite importar fichas familiares y personas de forma masiva.
"""
from datetime import datetime, date
from django.db import transaction
from django.core.exceptions import ValidationError
from openpyxl import load_workbook
import re
import logging
import os


# Configurar logger específico para importación
logger = logging.getLogger('importacion_masiva')


class ImportadorMasivo:
    """
    Clase para manejar la importación masiva de datos desde Excel.
    """

    def __init__(self, archivo_excel, organization):
        self.archivo = archivo_excel
        self.organization = organization
        self.errores = []
        self.advertencias = []
        self.fichas_creadas = 0
        self.personas_creadas = 0
        self.errores_detallados = []  # Lista de errores con contexto completo

        # Verificar si los datos de vivienda están habilitados
        self.housing_data_enabled = self._check_housing_data_enabled()

        # Configurar archivo de log
        self._setup_logging()

    def _check_housing_data_enabled(self):
        """Verifica si el parámetro de datos de vivienda está habilitado."""
        try:
            from censoapp.models import SystemParameters
            param = SystemParameters.objects.filter(key='Datos de Vivienda').first()
            return param and param.value == 'S'
        except Exception:
            # Si no existe el parámetro, asumimos que NO están habilitados
            return False

    def _setup_logging(self):
        """Configura el archivo de log para esta importación."""
        try:
            from django.conf import settings

            # Crear directorio de logs si no existe
            logs_dir = os.path.join(settings.MEDIA_ROOT, 'importacion_logs')
            os.makedirs(logs_dir, exist_ok=True)

            # Nombre del archivo de log con timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            org_name = self.organization.organization_name.replace(' ', '_')
            log_filename = f'importacion_{org_name}_{timestamp}.log'
            self.log_file_path = os.path.join(logs_dir, log_filename)

            # Configurar handler para este archivo
            file_handler = logging.FileHandler(self.log_file_path, encoding='utf-8')
            file_handler.setLevel(logging.INFO)

            # Formato detallado
            formatter = logging.Formatter(
                '%(asctime)s - %(levelname)s - %(message)s',
                datefmt='%Y-%m-%d %H:%M:%S'
            )
            file_handler.setFormatter(formatter)

            # Agregar handler al logger
            logger.addHandler(file_handler)

            # Log inicial
            logger.info('=' * 80)
            logger.info(f'INICIO DE IMPORTACIÓN MASIVA')
            logger.info(f'Organización: {self.organization.organization_name}')
            logger.info(f'Archivo: {os.path.basename(self.archivo)}')
            logger.info(f'Fecha/Hora: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
            logger.info('=' * 80)

        except Exception as e:
            # Si falla el setup de logging, no detener la importación
            self.log_file_path = None
            print(f"Advertencia: No se pudo configurar el archivo de log: {str(e)}")

    def _parsear_fecha(self, fecha_valor):
        """
        Parsea una fecha desde múltiples formatos comunes.
        Soporta: dd/mm/yyyy, dd-mm-yyyy, yyyy-mm-dd, y objetos datetime de Excel.
        Retorna un objeto date o None si no se puede parsear.
        """
        if not fecha_valor:
            return None

        # Si ya es un objeto date o datetime de Python
        if isinstance(fecha_valor, datetime):
            return fecha_valor.date()
        if isinstance(fecha_valor, date):
            return fecha_valor

        # Si es string, intentar diferentes formatos
        if isinstance(fecha_valor, str):
            fecha_str = fecha_valor.strip()

            # Lista de formatos a intentar (ordenados por más común)
            formatos = [
                '%d/%m/%Y',      # 25/12/2024
                '%d-%m-%Y',      # 25-12-2024
                '%Y-%m-%d',      # 2024-12-25 (ISO)
                '%d/%m/%y',      # 25/12/24
                '%d-%m-%y',      # 25-12-24
                '%Y/%m/%d',      # 2024/12/25
            ]

            for formato in formatos:
                try:
                    return datetime.strptime(fecha_str, formato).date()
                except ValueError:
                    continue

        # Si es un número (Excel a veces guarda fechas como números)
        if isinstance(fecha_valor, (int, float)):
            try:
                # Excel cuenta desde 1899-12-30
                from datetime import timedelta
                fecha_base = date(1899, 12, 30)
                return fecha_base + timedelta(days=int(fecha_valor))
            except:
                pass

        return None

    def validar_estructura(self):
        """Valida que el archivo tenga la estructura correcta."""
        try:
            wb = load_workbook(self.archivo, data_only=True)

            # Verificar que existan las hojas necesarias
            if 'Fichas' not in wb.sheetnames:
                self.errores.append("Falta la hoja 'Fichas' en el archivo Excel")
                return False

            if 'Personas' not in wb.sheetnames:
                self.errores.append("Falta la hoja 'Personas' en el archivo Excel")
                return False

            # Validar columnas de Fichas
            ws_fichas = wb['Fichas']
            columnas_fichas_requeridas = [
                'numero_ficha', 'vereda', 'zona', 'direccion'
            ]

            # Solo requerir columnas de materiales si el parámetro está habilitado
            if self.housing_data_enabled:
                columnas_fichas_requeridas.extend([
                    'material_paredes', 'material_piso', 'material_techo'
                ])

            headers_fichas = [cell.value for cell in ws_fichas[1]]
            for col in columnas_fichas_requeridas:
                if col not in headers_fichas:
                    self.errores.append(f"Falta la columna '{col}' en la hoja Fichas")

            # Validar columnas de Personas
            ws_personas = wb['Personas']
            columnas_personas_requeridas = [
                'numero_ficha', 'primer_nombre', 'primer_apellido',
                'identificacion', 'tipo_documento', 'fecha_nacimiento',
                'genero', 'parentesco', 'cabeza_familia'
            ]

            headers_personas = [cell.value for cell in ws_personas[1]]
            for col in columnas_personas_requeridas:
                if col not in headers_personas:
                    self.errores.append(f"Falta la columna '{col}' en la hoja Personas")

            return len(self.errores) == 0

        except FileNotFoundError:
            self.errores.append("El archivo Excel no fue encontrado")
            return False
        except PermissionError:
            self.errores.append("No se tienen permisos para leer el archivo Excel")
            return False
        except Exception as e:
            self.errores.append(f"Error al leer el archivo Excel: {str(e)}")
            return False

    def extraer_datos_fichas(self):
        """Extrae los datos de fichas del Excel."""
        try:
            wb = load_workbook(self.archivo, data_only=True)
            ws = wb['Fichas']

            # Obtener headers
            headers = [cell.value for cell in ws[1]]

            # Extraer datos
            fichas = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Saltar filas vacías
                    continue

                ficha_data = {}
                for idx, header in enumerate(headers):
                    if idx < len(row):  # Protección contra índices fuera de rango
                        ficha_data[header] = row[idx]

                ficha_data['fila_excel'] = row_idx
                fichas.append(ficha_data)

            return fichas
        except Exception as e:
            self.errores.append(f"Error al extraer datos de fichas: {str(e)}")
            return []

    def extraer_datos_personas(self):
        """Extrae los datos de personas del Excel."""
        try:
            wb = load_workbook(self.archivo, data_only=True)
            ws = wb['Personas']

            # Obtener headers
            headers = [cell.value for cell in ws[1]]

            # Extraer datos
            personas = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Saltar filas vacías
                    continue

                persona_data = {}
                for idx, header in enumerate(headers):
                    if idx < len(row):  # Protección contra índices fuera de rango
                        persona_data[header] = row[idx]

                persona_data['fila_excel'] = row_idx
                personas.append(persona_data)

            return personas
        except Exception as e:
            self.errores.append(f"Error al extraer datos de personas: {str(e)}")
            return []

    def validar_datos_ficha(self, ficha_data, fila):
        """Valida los datos de una ficha."""
        errores_fila = []

        # Validar número de ficha
        if not ficha_data.get('numero_ficha'):
            errores_fila.append(f"Fila {fila}: Número de ficha es obligatorio")

        # Validar vereda
        if not ficha_data.get('vereda'):
            errores_fila.append(f"Fila {fila}: Vereda es obligatoria")

        # Validar zona
        if ficha_data.get('zona') and ficha_data['zona'] not in ['U', 'R', 'URBANA', 'RURAL']:
            errores_fila.append(f"Fila {fila}: Zona debe ser 'U' (Urbana) o 'R' (Rural)")

        return errores_fila

    def validar_datos_persona(self, persona_data, fila):
        """Valida los datos de una persona."""
        errores_fila = []

        # Validar número de ficha
        if not persona_data.get('numero_ficha'):
            errores_fila.append(f"Fila {fila}: Número de ficha es obligatorio")

        # Validar nombres
        if not persona_data.get('primer_nombre'):
            errores_fila.append(f"Fila {fila}: Primer nombre es obligatorio")

        if not persona_data.get('primer_apellido'):
            errores_fila.append(f"Fila {fila}: Primer apellido es obligatorio")

        # Validar identificación
        if not persona_data.get('identificacion'):
            errores_fila.append(f"Fila {fila}: Identificación es obligatoria")
        else:
            # Verificar que sea alfanumérica
            if not re.match(r'^[A-Za-z0-9]+$', str(persona_data['identificacion'])):
                errores_fila.append(f"Fila {fila}: Identificación debe ser alfanumérica")

        # Validar fecha de nacimiento
        if not persona_data.get('fecha_nacimiento'):
            errores_fila.append(f"Fila {fila}: Fecha de nacimiento es obligatoria")
        else:
            # Intentar parsear la fecha en múltiples formatos
            fecha_parseada = self._parsear_fecha(persona_data['fecha_nacimiento'])
            if not fecha_parseada:
                errores_fila.append(
                    f"Fila {fila}: Fecha de nacimiento inválida. "
                    f"Formatos aceptados: dd/mm/yyyy, dd-mm-yyyy, yyyy-mm-dd"
                )
            else:
                # Validar que la fecha sea lógica (no en el futuro, no muy antigua)
                from datetime import date
                hoy = date.today()
                if fecha_parseada > hoy:
                    errores_fila.append(f"Fila {fila}: Fecha de nacimiento no puede ser futura")
                elif fecha_parseada < date(1900, 1, 1):
                    errores_fila.append(f"Fila {fila}: Fecha de nacimiento muy antigua (antes de 1900)")
                else:
                    # Guardar la fecha parseada para uso posterior
                    persona_data['fecha_nacimiento_parseada'] = fecha_parseada

        # Validar género
        if not persona_data.get('genero'):
            errores_fila.append(f"Fila {fila}: Género es obligatorio")

        # Validar cabeza de familia
        cabeza = str(persona_data.get('cabeza_familia', '')).upper()
        if cabeza not in ['SI', 'NO', 'SÍ', 'S', 'N', '1', '0', 'TRUE', 'FALSE']:
            errores_fila.append(f"Fila {fila}: Cabeza de familia debe ser SI/NO")

        return errores_fila

    def validar_duplicados(self, personas_data):
        """Valida que no haya identificaciones duplicadas."""
        from censoapp.models import Person

        identificaciones = [p.get('identificacion') for p in personas_data if p.get('identificacion')]

        # Duplicados en el Excel
        duplicados_excel = set([x for x in identificaciones if identificaciones.count(x) > 1])
        if duplicados_excel:
            for dup in duplicados_excel:
                self.errores.append(f"Identificación duplicada en Excel: {dup}")

        # Duplicados en la base de datos
        for persona_data in personas_data:
            identificacion = persona_data.get('identificacion')
            if identificacion:
                if Person.objects.filter(identification_person=identificacion).exists():
                    fila = persona_data.get('fila_excel', '?')
                    self.errores.append(f"Fila {fila}: Identificación {identificacion} ya existe en la base de datos")

    def validar_todo(self):
        """Ejecuta todas las validaciones."""
        # Validar estructura
        if not self.validar_estructura():
            return False

        # Extraer datos
        fichas_data = self.extraer_datos_fichas()
        personas_data = self.extraer_datos_personas()

        # Validar cada ficha
        for ficha in fichas_data:
            errores = self.validar_datos_ficha(ficha, ficha.get('fila_excel', '?'))
            self.errores.extend(errores)

        # Validar cada persona
        for persona in personas_data:
            errores = self.validar_datos_persona(persona, persona.get('fila_excel', '?'))
            self.errores.extend(errores)

        # Validar duplicados
        self.validar_duplicados(personas_data)

        # Validar que cada ficha tenga al menos una persona
        fichas_numeros = set([f.get('numero_ficha') for f in fichas_data])
        personas_fichas = set([p.get('numero_ficha') for p in personas_data])

        fichas_sin_personas = fichas_numeros - personas_fichas
        if fichas_sin_personas:
            for num_ficha in fichas_sin_personas:
                self.advertencias.append(f"Ficha {num_ficha} no tiene personas asociadas")

        return len(self.errores) == 0

    def importar(self):
        """
        Ejecuta la importación de datos.
        Retorna un diccionario con el resultado.
        """
        from censoapp.models import FamilyCard, Person, Sidewalks, Gender, Kinship
        from censoapp.models import IdentificationDocumentType, EducationLevel
        from censoapp.models import CivilState, Occupancy, SecuritySocial, Eps, Handicap
        from django.db import IntegrityError

        if not self.validar_todo():
            return {
                'exito': False,
                'errores': self.errores,
                'advertencias': self.advertencias
            }

        try:
            with transaction.atomic():
                # Extraer datos
                fichas_data = self.extraer_datos_fichas()
                personas_data = self.extraer_datos_personas()

                # Diccionario para mapear número de ficha a objeto FamilyCard
                fichas_creadas_map = {}

                # Crear fichas familiares
                for ficha_data in fichas_data:
                    # Obtener o crear vereda
                    vereda_nombre = ficha_data.get('vereda', 'Sin Vereda')
                    vereda, _ = Sidewalks.objects.get_or_create(
                        sidewalk_name=vereda_nombre,
                        organization_id=self.organization,
                        defaults={}
                    )

                    # Normalizar zona
                    zona = ficha_data.get('zona', 'R')
                    if zona in ['URBANA', 'U']:
                        zona = 'U'
                    else:
                        zona = 'R'

                    # Crear ficha familiar
                    ficha = FamilyCard.objects.create(
                        family_card_number=ficha_data.get('numero_ficha'),
                        sidewalk_home=vereda,
                        zone=zona,
                        address_home=ficha_data.get('direccion', ''),
                        organization=self.organization,
                        state=True
                    )

                    fichas_creadas_map[ficha_data.get('numero_ficha')] = ficha
                    self.fichas_creadas += 1

                # Crear personas
                for persona_data in personas_data:
                    try:
                        # Obtener la ficha
                        numero_ficha = persona_data.get('numero_ficha')
                        ficha = fichas_creadas_map.get(numero_ficha)

                        if not ficha:
                            error_msg = f"Ficha {numero_ficha} no encontrada para persona {persona_data.get('identificacion')}"
                            logger.warning(error_msg)
                            self.advertencias.append(error_msg)
                            continue

                        # Obtener o crear género
                        genero_nombre = persona_data.get('genero', 'Masculino')
                        genero, _ = Gender.objects.get_or_create(
                            gender=genero_nombre,
                            defaults={'gender_code': genero_nombre[0].upper()}
                        )

                        # Obtener o crear tipo de documento
                        tipo_doc_nombre = persona_data.get('tipo_documento', 'CC')
                        tipo_doc, _ = IdentificationDocumentType.objects.get_or_create(
                            code_document_type=tipo_doc_nombre,
                            defaults={'document_type': tipo_doc_nombre}
                        )

                        # Obtener o crear parentesco
                        parentesco_nombre = persona_data.get('parentesco', 'Otro')
                        parentesco, _ = Kinship.objects.get_or_create(
                            description_kinship=parentesco_nombre,
                            defaults={'code_kinship': '99'}
                        )

                        # Obtener defaults para campos obligatorios
                        edu_level = EducationLevel.objects.first()
                        civil_state = CivilState.objects.first()
                        occupation = Occupancy.objects.first()
                        security_social = SecuritySocial.objects.first()
                        eps = Eps.objects.first()
                        handicap = Handicap.objects.first()

                        # Normalizar cabeza de familia
                        cabeza_str = str(persona_data.get('cabeza_familia', 'NO')).upper()
                        es_cabeza = cabeza_str in ['SI', 'SÍ', 'S', '1', 'TRUE']

                        # Usar fecha parseada durante la validación o parsear ahora
                        fecha_nac = persona_data.get('fecha_nacimiento_parseada')
                        if not fecha_nac:
                            # Si no se parseó en validación, intentar ahora
                            fecha_nac = self._parsear_fecha(persona_data.get('fecha_nacimiento'))

                        if not fecha_nac:
                            # Usar fecha por defecto si no se pudo parsear
                            from datetime import date
                            fecha_nac = date(2000, 1, 1)

                        # Log de intento de creación
                        nombre_completo = f"{persona_data.get('primer_nombre', '')} {persona_data.get('primer_apellido', '')}"
                        logger.info(f"Creando persona: {nombre_completo} (ID: {persona_data.get('identificacion')}) - Ficha: {numero_ficha} - Cabeza: {es_cabeza}")

                        # Crear persona
                        persona = Person.objects.create(
                            first_name_1=persona_data.get('primer_nombre', ''),
                            first_name_2=persona_data.get('segundo_nombre', ''),
                            last_name_1=persona_data.get('primer_apellido', ''),
                            last_name_2=persona_data.get('segundo_apellido', ''),
                            identification_person=str(persona_data.get('identificacion', '')),
                            document_type=tipo_doc,
                            date_birth=fecha_nac,
                            gender=genero,
                            kinship=parentesco,
                            family_head=es_cabeza,
                            family_card=ficha,
                            education_level=edu_level,
                            civil_state=civil_state,
                            occupation=occupation,
                            social_insurance=security_social,
                            eps=eps,
                            handicap=handicap,
                            state=True
                        )

                        self.personas_creadas += 1
                        logger.info(f"✓ Persona creada exitosamente: {nombre_completo}")

                    except ValidationError as ve:
                        # Error de validación de Django
                        error_detail = {
                            'tipo': 'ValidationError',
                            'ficha': numero_ficha,
                            'persona': f"{persona_data.get('primer_nombre', '')} {persona_data.get('primer_apellido', '')}",
                            'identificacion': persona_data.get('identificacion'),
                            'error': str(ve.message_dict if hasattr(ve, 'message_dict') else ve),
                            'fila_excel': persona_data.get('fila_excel', 'N/A')
                        }

                        self.errores_detallados.append(error_detail)

                        # Log detallado
                        logger.error('=' * 60)
                        logger.error(f"ERROR DE VALIDACIÓN")
                        logger.error(f"Ficha: {numero_ficha}")
                        logger.error(f"Persona: {error_detail['persona']}")
                        logger.error(f"Identificación: {error_detail['identificacion']}")
                        logger.error(f"Fila Excel: {error_detail['fila_excel']}")
                        logger.error(f"Error: {error_detail['error']}")
                        logger.error('=' * 60)

                        # Agregar a errores generales
                        self.errores.append(
                            f"Ficha {numero_ficha} - {error_detail['persona']}: {error_detail['error']}"
                        )

                    except IntegrityError as ie:
                        # Error de integridad de base de datos
                        error_detail = {
                            'tipo': 'IntegrityError',
                            'ficha': numero_ficha,
                            'persona': f"{persona_data.get('primer_nombre', '')} {persona_data.get('primer_apellido', '')}",
                            'identificacion': persona_data.get('identificacion'),
                            'error': str(ie),
                            'fila_excel': persona_data.get('fila_excel', 'N/A')
                        }

                        self.errores_detallados.append(error_detail)

                        # Log detallado
                        logger.error('=' * 60)
                        logger.error(f"ERROR DE INTEGRIDAD")
                        logger.error(f"Ficha: {numero_ficha}")
                        logger.error(f"Persona: {error_detail['persona']}")
                        logger.error(f"Identificación: {error_detail['identificacion']}")
                        logger.error(f"Fila Excel: {error_detail['fila_excel']}")
                        logger.error(f"Error: {error_detail['error']}")
                        logger.error('=' * 60)

                        # Agregar a errores generales
                        self.errores.append(
                            f"Ficha {numero_ficha} - {error_detail['persona']}: Error de integridad - {str(ie)}"
                        )

                    except Exception as e:
                        # Cualquier otro error
                        error_detail = {
                            'tipo': 'Exception',
                            'ficha': numero_ficha,
                            'persona': f"{persona_data.get('primer_nombre', '')} {persona_data.get('primer_apellido', '')}",
                            'identificacion': persona_data.get('identificacion'),
                            'error': str(e),
                            'fila_excel': persona_data.get('fila_excel', 'N/A')
                        }

                        self.errores_detallados.append(error_detail)

                        # Log detallado
                        logger.error('=' * 60)
                        logger.error(f"ERROR INESPERADO")
                        logger.error(f"Ficha: {numero_ficha}")
                        logger.error(f"Persona: {error_detail['persona']}")
                        logger.error(f"Identificación: {error_detail['identificacion']}")
                        logger.error(f"Fila Excel: {error_detail['fila_excel']}")
                        logger.error(f"Error: {error_detail['error']}")
                        logger.error('=' * 60)

                        # Agregar a errores generales
                        self.errores.append(
                            f"Ficha {numero_ficha} - {error_detail['persona']}: {str(e)}"
                        )

                # Log de resumen
                logger.info('=' * 80)
                logger.info('RESUMEN DE IMPORTACIÓN')
                logger.info(f'Fichas creadas: {self.fichas_creadas}')
                logger.info(f'Personas creadas: {self.personas_creadas}')
                logger.info(f'Errores: {len(self.errores_detallados)}')
                logger.info(f'Advertencias: {len(self.advertencias)}')
                logger.info('=' * 80)

                if self.errores_detallados:
                    logger.info('ERRORES DETALLADOS:')
                    for error in self.errores_detallados:
                        logger.info(f"  - Ficha {error['ficha']}: {error['persona']} - {error['error']}")

                # Si hay errores, retornar con información del log
                if self.errores_detallados:
                    return {
                        'exito': False,
                        'fichas_creadas': self.fichas_creadas,
                        'personas_creadas': self.personas_creadas,
                        'errores': self.errores,
                        'errores_detallados': self.errores_detallados,
                        'advertencias': self.advertencias,
                        'log_file': self.log_file_path
                    }

                return {
                    'exito': True,
                    'fichas_creadas': self.fichas_creadas,
                    'personas_creadas': self.personas_creadas,
                    'advertencias': self.advertencias,
                    'log_file': self.log_file_path
                }

        except Exception as e:
            logger.error(f"ERROR CRÍTICO EN IMPORTACIÓN: {str(e)}", exc_info=True)
            return {
                'exito': False,
                'errores': [f"Error crítico durante la importación: {str(e)}"],
                'advertencias': self.advertencias,
                'log_file': self.log_file_path
            }

