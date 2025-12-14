from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import Paginator
from django.db import transaction, IntegrityError
from django.db.models import Value, Q, F, ExpressionWrapper, fields, Count, Sum
from django.db.models.functions.text import Concat
from django.http import HttpResponse, JsonResponse, Http404
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse, reverse_lazy
from django.utils.timezone import now
from django.views.decorators.http import require_POST
from django.views.generic import DetailView
from django.views.generic.edit import CreateView, UpdateView
from censoapp.models import Association, Person, FamilyCard, Sidewalks, SystemParameters, MaterialConstructionFamilyCard
from .forms import FormFamilyCard, FormPerson, MaterialConstructionFamilyForm
from django.views.decorators.csrf import csrf_protect
from .mixins import OrganizationFilterMixin, OrganizationPermissionMixin, OrganizationFormMixin, ReadOnlyPermissionMixin


# logger = logging.getLogger(__name__)


# Create your views here.
@login_required
def home(request):
    personas = Person.objects.values(
        'gender__gender',
        'date_birth',
        'family_card_id',
        'family_card__sidewalk_home__sidewalk_name'
    ).annotate(
        total=Count('id')
    )

    # Obtener la cantidad de personas por vereda
    personas_veredas = Person.objects.values('family_card__sidewalk_home__sidewalk_name').annotate(
        total_personas=Count('id')
    ).order_by('family_card__sidewalk_home__sidewalk_name')

    # Crear un diccionario para almacenar los totales por vereda
    veredas_totales = {}
    for vereda in personas_veredas:
        vereda_nombre = vereda['family_card__sidewalk_home__sidewalk_name']
        veredas_totales[vereda_nombre] = vereda['total_personas']

    veredas_totales = dict(sorted(veredas_totales.items(), key=lambda item: item[0],  reverse=True))


    total_personas = Person.objects.count()
    total_fichas = FamilyCard.objects.count()
    total_veredas = Sidewalks.objects.count()

    # Obtener el total de personas por género y por año de nacimiento
    personas_por_genero = Person.objects.values('gender__gender', 'date_birth').annotate(personas=Count('id'))

    mujeres = {}
    hombres = {}

    for genero in personas_por_genero:
        if genero['gender__gender'] == 'Femenino':
            anio = genero['date_birth'].year
            mujeres[anio] = mujeres.get(anio, 0) + genero['personas']
        elif genero['gender__gender'] == 'Masculino':
            anio = genero['date_birth'].year
            hombres[anio] = hombres.get(anio, 0) + genero['personas']

    # Crear un diccionario para almacenar los datos de los años
    anios = sorted(set(list(mujeres.keys()) + list(hombres.keys())))
    mujeres_list = [mujeres.get(anio, 0) for anio in anios]
    hombres_list = [hombres.get(anio, 0) for anio in anios]


    context = {
        'segment': 'dashboard',
        'total_personas': total_personas,
        'total_fichas': total_fichas,
        'total_veredas': total_veredas,
        'hombres_list': hombres_list,
        'mujeres_list': mujeres_list,
        'anios': anios,
        'veredas_totales': veredas_totales,
    }
    return render(request, 'censo/dashboard.html', context)


@login_required
def profile(request):
    return render(request, 'account/profile.html')


@login_required
def association(request):
    """
    Vista optimizada para mostrar asociaciones.
    Incluye paginación, búsqueda y carga optimizada de datos.
    """
    try:
        # Query optimizado
        associations_list = Association.objects.all().order_by('-id')

        # Búsqueda (si se proporciona)
        search_query = request.GET.get('search', '').strip()
        if search_query:
            associations_list = associations_list.filter(
                Q(association_name__icontains=search_query) |
                Q(association_identification__icontains=search_query) |
                Q(association_email__icontains=search_query)
            )

        # Paginación
        paginator = Paginator(associations_list, 10)  # 10 por página
        page_number = request.GET.get('page', 1)

        try:
            associations = paginator.get_page(page_number)
        except Exception:
            associations = paginator.get_page(1)

        context = {
            'associations': associations,
            'search_query': search_query,
            'total_associations': associations_list.count(),
            'segment': 'association'
        }

        return render(request, 'censo/configuracion/association.html', context)

    except Exception as e:
        messages.error(request, f"Error al cargar las asociaciones: {str(e)}")
        return render(request, 'censo/configuracion/association.html', {
            'associations': [],
            'segment': 'association'
        })


class CreateAssociation(CreateView):
    model = Association
    fields = '__all__'
    template_name = 'censo/createAssociation.html'
    success_url = reverse_lazy('association')

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super(CreateAssociation, self).form_valid(form)


def family_card_index(request):
    return render(request, 'censo/censo/familyCardIndex.html',
                  {'segment': 'family_card'})


# Función optimizada para obtener las fichas familiares en formato JSON para DataTables
@login_required
def get_family_cards(request):
    """
    Vista optimizada para listar fichas familiares con paginación server-side.
    Incluye búsqueda multi-campo, ordenamiento y conteo de miembros.
    """
    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '').strip()
        order_column = request.GET.get('order[0][column]', '0')
        order_dir = request.GET.get('order[0][dir]', 'asc')

        # Columnas ordenables
        order_columns = [
            'family_card__family_card_number',
            'full_name',
            'family_card__sidewalk_home__sidewalk_name',
            'person_count',
        ]

        # Validar índice de columna
        try:
            order_by = order_columns[int(order_column)]
            if order_dir == 'desc':
                order_by = f'-{order_by}'
        except (IndexError, ValueError):
            order_by = 'family_card__family_card_number'

        # Query optimizado con select_related para evitar N+1 queries
        queryset = (Person.objects
                    .select_related('family_card', 'family_card__sidewalk_home',
                                  'family_card__organization', 'document_type')
                    .filter(family_head=True, state=True, family_card__state=True))

        # Filtrar por organización del usuario (multi-tenancy)
        if not (request.user.is_superuser or getattr(request, 'can_view_all', False)):
            user_organization = getattr(request, 'user_organization', None)
            if user_organization:
                queryset = queryset.filter(family_card__organization=user_organization)
            else:
                # Usuario sin organización, retornar vacío
                return JsonResponse({
                    'draw': draw,
                    'recordsTotal': 0,
                    'recordsFiltered': 0,
                    'data': []
                })

        queryset = queryset.values('family_card__family_card_number', 'family_card_id',
                           'first_name_1', 'first_name_2', 'last_name_1', 'last_name_2',
                           'identification_person', 'document_type__code_document_type',
                           'family_card__sidewalk_home__sidewalk_name', 'family_card__zone',
                           'family_card__address_home', 'family_card__organization__organization_name').annotate(
                        full_name=Concat('first_name_1', Value(' '), 'first_name_2',
                                       Value(' '), 'last_name_1', Value(' '), 'last_name_2'),
                        person_count=Count('family_card__person',
                                         filter=Q(family_card__person__state=True))
                    )

        # Búsqueda optimizada con OR en múltiples campos
        if search_value:
            queryset = queryset.filter(
                Q(first_name_1__icontains=search_value) |
                Q(first_name_2__icontains=search_value) |
                Q(last_name_1__icontains=search_value) |
                Q(last_name_2__icontains=search_value) |
                Q(identification_person__icontains=search_value) |
                Q(family_card__family_card_number__icontains=search_value) |
                Q(family_card__sidewalk_home__sidewalk_name__icontains=search_value) |
                Q(family_card__zone__icontains=search_value) |
                Q(family_card__address_home__icontains=search_value)
            )

        # Total de registros antes de aplicar filtros (respetando organización)
        total_queryset = Person.objects.filter(
            family_head=True,
            state=True,
            family_card__state=True
        )

        # Aplicar filtro de organización al total también
        if not (request.user.is_superuser or getattr(request, 'can_view_all', False)):
            user_organization = getattr(request, 'user_organization', None)
            if user_organization:
                total_queryset = total_queryset.filter(family_card__organization=user_organization)

        total_records = total_queryset.count()

        # Total de registros después de aplicar filtros
        filtered_records = queryset.count()

        # Aplicar ordenamiento
        queryset = queryset.order_by(order_by)

        # Paginación
        paginator = Paginator(queryset, length)
        page_number = (start // length) + 1

        try:
            page = paginator.get_page(page_number)
        except Exception:
            page = paginator.get_page(1)

        # Serializar los datos
        data = list(page.object_list)

        # Respuesta JSON para DataTables
        response_data = {
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': filtered_records,
            'data': data
        }

        return JsonResponse(response_data)

    except Exception as e:
        # logger.error(f"Error en get_family_cards: {e}")
        return JsonResponse({
            'draw': 1,
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': 'Error al cargar los datos'
        }, status=500)


# Función optimizada para crear ficha familiar con cabeza de familia
@login_required
def create_family_card(request):
    """
    Vista para crear una nueva ficha familiar junto con el cabeza de familia.
    Incluye validaciones robustas y mensajes claros para el usuario.
    Valida permisos de escritura (usuarios VIEWER no pueden crear).
    """
    # Validar permisos de escritura
    if not request.user.is_superuser:
        user_role = getattr(request, 'user_role', None)
        if user_role == 'VIEWER':
            messages.error(
                request,
                "No tiene permisos para crear fichas familiares. Su rol es de solo lectura."
            )
            return redirect('familyCardIndex')

    if request.method == 'POST':
        family_card_form = FormFamilyCard(request.POST)
        person_form = FormPerson(request.POST)

        # Validación 1: Verificar duplicidad de identificación ANTES de validar formularios
        identification_person = request.POST.get('identification_person', '').strip()
        if identification_person:
            existing_person = Person.objects.filter(
                identification_person=identification_person,
                state=True
            ).first()

            if existing_person:
                messages.error(
                    request,
                    f"El documento {identification_person} ya está registrado para "
                    f"{existing_person.full_name} en la ficha {existing_person.family_card.family_card_number}."
                )
                return render(request, 'censo/censo/createFamilyCard.html', {
                    'family_card_form': family_card_form,
                    'person_form': person_form,
                    'segment': 'family_card'
                })

        # Validación 2: Validar edad mínima del cabeza de familia (18 años)
        date_birth = request.POST.get('date_birth')
        if date_birth:
            from datetime import date, datetime
            try:
                birth_date = datetime.strptime(date_birth, '%Y-%m-%d').date()
                today = date.today()
                age = today.year - birth_date.year - (
                    (today.month, today.day) < (birth_date.month, birth_date.day)
                )

                if age < 18:
                    messages.error(
                        request,
                        f"El cabeza de familia debe ser mayor de 18 años. "
                        f"La edad registrada es de {age} años."
                    )
                    return render(request, 'censo/censo/createFamilyCard.html', {
                        'family_card_form': family_card_form,
                        'person_form': person_form,
                        'segment': 'family_card'
                    })
            except ValueError:
                messages.error(request, "Formato de fecha de nacimiento inválido.")
                return render(request, 'censo/censo/createFamilyCard.html', {
                    'family_card_form': family_card_form,
                    'person_form': person_form,
                    'segment': 'family_card'
                })

        # Validación 3: Validar formularios
        if family_card_form.is_valid() and person_form.is_valid():
            try:
                with transaction.atomic():
                    # Crear FamilyCard
                    family_card = family_card_form.save(commit=False)
                    family_card.family_card_number = FamilyCard.get_next_family_card_number()
                    family_card.state = True
                    family_card.save()

                    # Crear Person (cabeza de familia)
                    person = person_form.save(commit=False)
                    person.family_card = family_card
                    person.family_head = True
                    person.state = True
                    person.save()

                    messages.success(
                        request,
                        f"Ficha familiar #{family_card.family_card_number} creada exitosamente. "
                        f"Cabeza de familia: {person.full_name}"
                    )
                    return redirect('createPerson', pk=family_card.pk)

            except IntegrityError as e:
                messages.error(
                    request,
                    "Error al guardar: Ya existe un registro con estos datos. "
                    "Por favor, verifique la información."
                )
            except Exception as e:
                messages.error(
                    request,
                    "Ocurrió un error inesperado al crear la ficha familiar. "
                    "Por favor, contacte al administrador."
                )
        else:
            # Mostrar errores específicos de los formularios
            if family_card_form.errors:
                for field, errors in family_card_form.errors.items():
                    for error in errors:
                        messages.error(request, f"Datos de vivienda - {field}: {error}")

            if person_form.errors:
                for field, errors in person_form.errors.items():
                    for error in errors:
                        messages.error(request, f"Datos de persona - {field}: {error}")

            if not (family_card_form.errors or person_form.errors):
                messages.warning(
                    request,
                    "Por favor, complete todos los campos obligatorios del formulario."
                )
    else:
        family_card_form = FormFamilyCard()
        person_form = FormPerson()

    return render(request, 'censo/censo/createFamilyCard.html', {
        'family_card_form': family_card_form,
        'person_form': person_form,
        'segment': 'family_card'
    })


@login_required
def crear_persona(request, pk):
    """
    Vista para agregar un nuevo miembro a una ficha familiar existente.
    Incluye validaciones robustas y mensajes claros para el usuario.
    Valida permisos de escritura (usuarios VIEWER no pueden crear).
    """
    # Validar permisos de escritura
    if not request.user.is_superuser:
        user_role = getattr(request, 'user_role', None)
        if user_role == 'VIEWER':
            messages.error(
                request,
                "No tiene permisos para crear personas. Su rol es de solo lectura."
            )
            return redirect('familyCardIndex')

    # Validar que la ficha familiar existe y está activa
    try:
        familia = get_object_or_404(FamilyCard, pk=pk, state=True)
    except Exception:
        messages.error(request, "La ficha familiar no existe o no está activa.")
        return redirect('familyCardIndex')

    if request.method == 'POST':
        person_form = FormPerson(request.POST)

        # Validación 1: Verificar duplicidad de identificación ANTES de validar formulario
        identification_person = request.POST.get('identification_person', '').strip()
        if identification_person:
            existing_person = Person.objects.filter(
                identification_person=identification_person,
                state=True
            ).first()

            if existing_person:
                messages.error(
                    request,
                    f"El documento {identification_person} ya está registrado para "
                    f"{existing_person.full_name} en la ficha #{existing_person.family_card.family_card_number}."
                )
                return render(request, 'censo/persona/createPerson.html', {
                    'person_form': person_form,
                    'familia': familia,
                    'segment': 'family_card'
                })

        # Validación 2: Si se marca como cabeza de familia, validar que sea mayor de 18 años
        is_family_head = request.POST.get('family_head') == 'on'
        date_birth = request.POST.get('date_birth')

        if is_family_head and date_birth:
            from datetime import date, datetime
            try:
                birth_date = datetime.strptime(date_birth, '%Y-%m-%d').date()
                today = date.today()
                age = today.year - birth_date.year - (
                    (today.month, today.day) < (birth_date.month, birth_date.day)
                )

                if age < 18:
                    messages.error(
                        request,
                        f"El cabeza de familia debe ser mayor de 18 años. "
                        f"La edad registrada es de {age} años."
                    )
                    return render(request, 'censo/persona/createPerson.html', {
                        'person_form': person_form,
                        'familia': familia,
                        'segment': 'family_card'
                    })
            except ValueError:
                messages.error(request, "Formato de fecha de nacimiento inválido.")
                return render(request, 'censo/persona/createPerson.html', {
                    'person_form': person_form,
                    'familia': familia,
                    'segment': 'family_card'
                })

        # Validación 3: Verificar que no haya otro cabeza de familia si se marca como tal
        if is_family_head:
            existing_head = Person.objects.filter(
                family_card=familia,
                family_head=True,
                state=True
            ).first()

            if existing_head:
                messages.error(
                    request,
                    f"Ya existe un cabeza de familia en esta ficha: {existing_head.full_name}. "
                    f"Primero debe cambiar el rol del cabeza actual."
                )
                return render(request, 'censo/persona/createPerson.html', {
                    'person_form': person_form,
                    'familia': familia,
                    'segment': 'family_card'
                })

        # Validación 4: Validar formulario
        if person_form.is_valid():
            try:
                with transaction.atomic():
                    person = person_form.save(commit=False)
                    person.family_card = familia
                    person.state = True
                    person.save()

                    # Mensaje de éxito personalizado
                    messages.success(
                        request,
                        f"Miembro agregado exitosamente: {person.full_name}. "
                        f"Total de miembros en la ficha: {familia.person_set.filter(state=True).count()}"
                    )

                    # Redireccionar según la acción del usuario
                    if 'add_another' in request.POST:
                        return redirect('createPerson', pk=familia.pk)
                    else:
                        return redirect('detailFamilyCard', pk=familia.pk)

            except IntegrityError:
                messages.error(
                    request,
                    "Error al guardar: Ya existe un registro con estos datos. "
                    "Por favor, verifique la información."
                )
            except Exception:
                messages.error(
                    request,
                    "Ocurrió un error inesperado al agregar el miembro. "
                    "Por favor, contacte al administrador."
                )
        else:
            # Mostrar errores específicos del formulario
            if person_form.errors:
                for field, errors in person_form.errors.items():
                    for error in errors:
                        field_label = person_form.fields.get(field).label if field in person_form.fields else field
                        messages.error(request, f"{field_label}: {error}")

            if not person_form.errors:
                messages.warning(
                    request,
                    "Por favor, complete todos los campos obligatorios del formulario."
                )
    else:
        person_form = FormPerson()

    # Obtener información adicional de la familia para el contexto
    total_members = familia.person_set.filter(state=True).count()
    family_head = familia.person_set.filter(family_head=True, state=True).first()

    return render(request, 'censo/persona/createPerson.html', {
        'person_form': person_form,
        'familia': familia,
        'total_members': total_members,
        'family_head': family_head,
        'segment': 'family_card'
    })


# Muestra el detalle de la ficha familiar - Vista optimizada
@login_required
def detalle_ficha(request, pk):
    """
    Vista optimizada para mostrar el detalle de una ficha familiar.
    Incluye información de la vivienda y todos los miembros de la familia.
    Valida permisos de organización (multi-tenancy).
    """
    try:
        # Verificar que la ficha familiar existe
        family_card = get_object_or_404(FamilyCard, pk=pk, state=True)

        # Validar permisos de organización
        if not (request.user.is_superuser or getattr(request, 'can_view_all', False)):
            user_organization = getattr(request, 'user_organization', None)
            if not user_organization or family_card.organization != user_organization:
                messages.error(request, "No tiene permiso para acceder a esta ficha familiar.")
                return redirect('familyCardIndex')

        # Query optimizado con select_related para evitar N+1 queries
        familia = (Person.objects
                   .select_related('family_card', 'family_card__sidewalk_home',
                                 'family_card__organization', 'kinship', 'document_type', 'gender')
                   .filter(family_card_id=pk, state=True)
                   .values('id', 'first_name_1', 'first_name_2', 'last_name_1', 'last_name_2',
                          'date_birth', 'identification_person', 'document_type__code_document_type',
                          'kinship__description_kinship', 'family_card__family_card_number',
                          'family_card__sidewalk_home__sidewalk_name', 'family_head',
                          'family_card__zone', 'family_card__address_home', 'family_card__id',
                          'family_card__latitude', 'family_card__longitude',
                          'family_card__organization__organization_name', 'gender__gender',
                          'cell_phone', 'personal_email')
                   .annotate(
                       age=ExpressionWrapper(now().year - F('date_birth__year'),
                                           output_field=fields.IntegerField())
                   )
                   .order_by('-family_head', 'date_birth'))

        # Verificar que hay miembros en la familia
        if not familia.exists():
            messages.warning(request, "No se encontraron miembros en esta ficha familiar.")

        # Calcular estadísticas de la familia
        total_miembros = familia.count()
        cabeza_familia = familia.filter(family_head=True).first()
        promedio_edad = familia.aggregate(promedio=ExpressionWrapper(
            Sum(now().year - F('date_birth__year')) / Count('id'),
            output_field=fields.FloatField()
        ))['promedio']

        context = {
            'familia': familia,
            'family_card': family_card,
            'family_card_obj': family_card,  # Para acceder al historial
            'total_miembros': total_miembros,
            'cabeza_familia': cabeza_familia,
            'promedio_edad': round(promedio_edad, 1) if promedio_edad else 0,
            'segment': 'family_card',
        }

        return render(request, 'censo/censo/detail_family_card.html', context)

    except Exception as e:
        messages.error(request, "Hubo un error al cargar el detalle de la ficha familiar.")
        return redirect('familyCardIndex')


class UpdateFamily(LoginRequiredMixin, ReadOnlyPermissionMixin, OrganizationPermissionMixin,
                   OrganizationFilterMixin, OrganizationFormMixin, UpdateView):
    """
    Vista optimizada para actualizar fichas familiares.
    Maneja dos formularios: datos de ubicación y datos de vivienda.
    Incluye validaciones robustas y mensajes claros para el usuario.

    Mixins:
    - ReadOnlyPermissionMixin: Bloquea acceso a usuarios VIEWER
    - OrganizationPermissionMixin: Valida que el usuario tenga permiso para editar la ficha
    - OrganizationFilterMixin: Filtra fichas por organización del usuario
    - OrganizationFormMixin: Limita opciones de organización/vereda en formularios
    """
    model = FamilyCard
    form_class = FormFamilyCard
    template_name = 'censo/censo/edit-family-card.html'
    success_url = reverse_lazy('familyCardIndex')

    def get_queryset(self):
        """Optimizar query con select_related para evitar N+1"""
        # El mixin ya filtra por organización, aquí solo optimizamos
        queryset = super().get_queryset()
        return queryset.select_related(
            'sidewalk_home',
            'organization'
        ).filter(state=True)

    def form_valid(self, form):
        """Validar y guardar el formulario de ubicación"""
        try:
            # Validar coordenadas si se proporcionan
            latitude = form.cleaned_data.get('latitude')
            longitude = form.cleaned_data.get('longitude')

            if latitude or longitude:
                if not (latitude and longitude):
                    messages.error(
                        self.request,
                        "Debe proporcionar tanto la latitud como la longitud, o dejar ambas en blanco."
                    )
                    return self.form_invalid(form)

                # Validar rangos
                try:
                    lat_float = float(latitude) if isinstance(latitude, str) else latitude
                    lon_float = float(longitude) if isinstance(longitude, str) else longitude

                    if not (-90 <= lat_float <= 90):
                        messages.error(
                            self.request,
                            f"La latitud debe estar entre -90 y 90 grados. Valor ingresado: {lat_float}"
                        )
                        return self.form_invalid(form)

                    if not (-180 <= lon_float <= 180):
                        messages.error(
                            self.request,
                            f"La longitud debe estar entre -180 y 180 grados. Valor ingresado: {lon_float}"
                        )
                        return self.form_invalid(form)
                except (ValueError, TypeError):
                    messages.error(
                        self.request,
                        "Las coordenadas deben ser valores numéricos válidos."
                    )
                    return self.form_invalid(form)

            # Guardar
            with transaction.atomic():
                form.instance.user = self.request.user
                response = super(UpdateFamily, self).form_valid(form)

                messages.success(
                    self.request,
                    f"Ficha familiar #{self.object.family_card_number} actualizada correctamente."
                )
                return response

        except Exception as e:
            messages.error(
                self.request,
                "Ocurrió un error al actualizar la ficha familiar. Por favor, intente nuevamente."
            )
            return self.form_invalid(form)

    def form_invalid(self, form):
        """Mostrar errores específicos del formulario"""
        if form.errors:
            for field, errors in form.errors.items():
                for error in errors:
                    field_name = form.fields[field].label if field in form.fields else field
                    messages.error(self.request, f"{field_name}: {error}")
        else:
            messages.warning(
                self.request,
                "Hubo un problema con la actualización. Por favor, revise los campos."
            )
        return super(UpdateFamily, self).form_invalid(form)

    def post(self, request, *args, **kwargs):
        """
        Procesa envíos del formulario de vivienda o del formulario principal.
        Detecta automáticamente cuál formulario se está enviando.
        """
        self.object = self.get_object()

        # Detectar si es el formulario de materiales de vivienda
        if 'material_form_submit' in request.POST:
            return self._handle_material_form(request)

        # Formulario principal de ubicación
        return super().post(request, *args, **kwargs)

    def _handle_material_form(self, request):
        """Procesar el formulario de materiales de vivienda de manera optimizada"""
        try:
            # Obtener instancia existente o None
            material_instance = MaterialConstructionFamilyCard.get_materials_by_family_card(
                self.object.pk
            )

            # Determinar acción según si existe o no
            action = "actualizados" if material_instance else "guardados"

            # Crear formulario con instancia (existente o None)
            material_form = MaterialConstructionFamilyForm(
                request.POST,
                instance=material_instance
            )

            if material_form.is_valid():
                try:
                    with transaction.atomic():
                        instance = material_form.save(commit=False)
                        instance.family_card = self.object
                        instance.save()

                        # Guardar relaciones many-to-many si existen
                        if hasattr(material_form, 'save_m2m'):
                            material_form.save_m2m()

                        messages.success(
                            request,
                            f"✓ Datos de vivienda {action} correctamente para la ficha #{self.object.family_card_number}."
                        )

                        # Redirigir y mantener la pestaña vivienda activa
                        url = reverse('update-family', kwargs={'pk': self.object.pk}) + '?tab=vivienda'
                        return redirect(url)

                except IntegrityError as e:
                    messages.error(
                        request,
                        "Error de integridad: Ya existe un registro de vivienda para esta ficha."
                    )
                except Exception as e:
                    messages.error(
                        request,
                        f"Error al guardar los datos de vivienda. Por favor, intente nuevamente."
                    )
            else:
                # Mostrar errores específicos del formulario
                for field, errors in material_form.errors.items():
                    for error in errors:
                        if field == '__all__':
                            messages.error(request, f"Error: {error}")
                        else:
                            field_name = material_form.fields.get(field).label if field in material_form.fields else field
                            messages.error(request, f"Vivienda - {field_name}: {error}")

            # Renderizar con errores
            return self.render_to_response(
                self.get_context_data(
                    form=self.get_form(),
                    material_form=material_form
                )
            )

        except Exception as e:
            messages.error(
                request,
                "Error inesperado al procesar el formulario de vivienda."
            )
            return redirect('update-family', pk=self.object.pk)

    def get_context_data(self, **kwargs):
        """Agregar contexto optimizado con información de la ficha"""
        context = super().get_context_data(**kwargs)

        # Parámetros del sistema (cacheados si es posible)
        params = SystemParameters.objects.all().only('key', 'value')
        system_params = {p.key: p.value for p in params}

        context['segment'] = 'family_card'
        context['system_params'] = system_params
        context['datos_vivienda'] = system_params.get('Datos de Vivienda', 'N')

        # Información de la familia para el contexto
        family = self.object

        # Contar miembros activos de forma eficiente
        context['total_members'] = Person.objects.filter(
            family_card=family,
            state=True
        ).count()

        # Obtener cabeza de familia
        context['family_head'] = Person.objects.filter(
            family_card=family,
            family_head=True,
            state=True
        ).only('first_name_1', 'last_name_1').first()

        # Formulario de materiales de vivienda
        material_instance = None
        try:
            material_instance = MaterialConstructionFamilyCard.get_materials_by_family_card(family.pk)
        except Exception:
            pass

        if 'material_form' not in kwargs:
            if material_instance:
                context['material_form'] = MaterialConstructionFamilyForm(instance=material_instance)
                context['material_exists'] = True
            else:
                context['material_form'] = MaterialConstructionFamilyForm()
                context['material_exists'] = False
        else:
            context['material_form'] = kwargs['material_form']
            context['material_exists'] = bool(material_instance)

        return context


class DetailPersona(LoginRequiredMixin, OrganizationPermissionMixin,
                    OrganizationFilterMixin, DetailView):
    """
    Vista optimizada para mostrar el detalle de una persona.
    Incluye validaciones, carga optimizada de datos relacionados y contexto enriquecido.

    Mixins:
    - OrganizationPermissionMixin: Valida que el usuario tenga permiso para ver la persona
    - OrganizationFilterMixin: Filtra personas por organización del usuario
    """
    model = Person
    template_name = 'censo/persona/detail_person.html'
    context_object_name = 'persona'

    def get_queryset(self):
        """
        Query optimizado con select_related para evitar N+1 queries.
        Solo carga registros activos para seguridad.
        El mixin ya filtra por organización.
        """
        queryset = super().get_queryset()
        return queryset.select_related(
            'document_type',
            'gender',
            'education_level',
            'civil_state',
            'occupation',
            'social_insurance',
            'eps',
            'kinship',
            'family_card',
            'family_card__sidewalk_home',
            'family_card__organization',
            'handicap'
        ).filter(state=True)

    def get_object(self, queryset=None):
        """Validar que la persona existe y está activa"""
        try:
            obj = super().get_object(queryset)
            if not obj.state:
                raise Http404("Esta persona no está disponible.")
            return obj
        except Person.DoesNotExist:
            raise Http404("Persona no encontrada.")

    def get_context_data(self, **kwargs):
        """Agregar contexto enriquecido con información relacionada"""
        context = super().get_context_data(**kwargs)
        persona = self.object

        # Calcular edad
        if persona.date_birth:
            from datetime import date
            today = date.today()
            age = today.year - persona.date_birth.year - (
                (today.month, today.day) < (persona.date_birth.month, persona.date_birth.day)
            )
            context['age'] = age
        else:
            context['age'] = None

        # Información de la familia (optimizado)
        if persona.family_card:
            family = persona.family_card

            # Total de miembros de la familia
            context['total_family_members'] = Person.objects.filter(
                family_card=family,
                state=True
            ).count()

            # Cabeza de familia
            context['family_head_obj'] = Person.objects.filter(
                family_card=family,
                family_head=True,
                state=True
            ).only('first_name_1', 'last_name_1', 'id').first()

            # Miembros de la familia (limitado a 10 para rendimiento)
            context['family_members'] = Person.objects.filter(
                family_card=family,
                state=True
            ).exclude(
                id=persona.id
            ).select_related(
                'kinship', 'gender'
            ).only(
                'id', 'first_name_1', 'last_name_1', 'kinship__description_kinship',
                'gender__gender', 'date_birth', 'family_head'
            )[:10]

            context['family_card'] = family
        else:
            context['total_family_members'] = 0
            context['family_head_obj'] = None
            context['family_members'] = []
            context['family_card'] = None

        # Información de seguridad
        context['segment'] = 'persons'
        context['can_edit'] = self.request.user.is_authenticated

        return context



# Lista las personas en formato JSON para DataTables
@login_required
def listar_personas(request):
    """
    Vista optimizada para listar personas con paginación server-side.
    Incluye búsqueda, ordenamiento y cálculo de edad eficiente.
    """
    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '').strip()
        order_column = request.GET.get('order[0][column]', '0')
        order_dir = request.GET.get('order[0][dir]', 'asc')

        # Columnas ordenables
        order_columns = [
            'first_name_1',
            'date_birth',
            'age',
            'gender__gender',
            'family_card__family_card_number',
        ]

        # Validar índice de columna
        try:
            order_by = order_columns[int(order_column)]
            if order_dir == 'desc':
                order_by = f'-{order_by}'
        except (IndexError, ValueError):
            order_by = 'first_name_1'

        # Query optimizado con select_related para evitar N+1 queries
        personas = (Person.objects
                    .select_related('document_type', 'gender', 'family_card', 'family_card__sidewalk_home',
                                  'family_card__organization')
                    .filter(state=True))

        # Filtrar por organización del usuario (multi-tenancy)
        if not (request.user.is_superuser or getattr(request, 'can_view_all', False)):
            user_organization = getattr(request, 'user_organization', None)
            if user_organization:
                personas = personas.filter(family_card__organization=user_organization)
            else:
                # Usuario sin organización, retornar vacío
                return JsonResponse({
                    'draw': draw,
                    'recordsTotal': 0,
                    'recordsFiltered': 0,
                    'data': []
                })

        personas = personas.values('id', 'first_name_1', 'first_name_2', 'last_name_1', 'last_name_2',
                            'identification_person', 'document_type__code_document_type', 'date_birth',
                            'family_card', 'family_head', 'family_card__family_card_number',
                            'family_card__sidewalk_home__sidewalk_name').annotate(
                        gender=F('gender__gender'),
                        age=ExpressionWrapper(
                            now().year - F('date_birth__year'),
                            output_field=fields.IntegerField()
                        )
                    )

        # Búsqueda optimizada con OR en múltiples campos
        if search_value:
            personas = personas.filter(
                Q(first_name_1__icontains=search_value) |
                Q(first_name_2__icontains=search_value) |
                Q(last_name_1__icontains=search_value) |
                Q(last_name_2__icontains=search_value) |
                Q(identification_person__icontains=search_value) |
                Q(family_card__family_card_number__icontains=search_value) |
                Q(family_card__sidewalk_home__sidewalk_name__icontains=search_value)
            )

        # Total de registros antes de aplicar filtros (para DataTables)
        # Respetar filtro de organización también en el total
        total_queryset = Person.objects.filter(state=True)

        if not (request.user.is_superuser or getattr(request, 'can_view_all', False)):
            user_organization = getattr(request, 'user_organization', None)
            if user_organization:
                total_queryset = total_queryset.filter(family_card__organization=user_organization)

        total_records = total_queryset.count()

        # Total de registros después de aplicar filtros
        filtered_records = personas.count()

        # Aplicar ordenamiento
        personas = personas.order_by(order_by)

        # Paginación
        paginator = Paginator(personas, length)
        page_number = (start // length) + 1

        try:
            page = paginator.get_page(page_number)
        except Exception:
            page = paginator.get_page(1)

        # Serializar los datos
        data = list(page.object_list)

        # Respuesta JSON para DataTables
        response_data = {
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': filtered_records,
            'data': data
        }

        return JsonResponse(response_data)

    except Exception as e:
        # logger.error(f"Error en listar_personas: {e}")
        return JsonResponse({
            'draw': 1,
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': 'Error al cargar los datos'
        }, status=500)


# Vista para mostrar la lista de personas en la plantilla HTML.
def view_persons(request):
    return render(request, 'censo/persona/listado_personas.html', {'segment': 'personas'})


# Vista para editar una persona
class UpdatePerson(LoginRequiredMixin, ReadOnlyPermissionMixin, OrganizationPermissionMixin,
                   OrganizationFilterMixin, UpdateView):
    """
    Vista para editar una persona.

    Mixins:
    - ReadOnlyPermissionMixin: Bloquea acceso a usuarios VIEWER
    - OrganizationPermissionMixin: Valida que el usuario tenga permiso para editar
    - OrganizationFilterMixin: Filtra personas por organización del usuario
    """
    model = Person
    fields = ['first_name_1', 'first_name_2', 'last_name_1', 'last_name_2', 'document_type', 'identification_person',
              'date_birth', 'cell_phone', 'personal_email', 'gender', 'kinship', 'education_level', 'civil_state',
              'occupation', 'social_insurance', 'eps', 'handicap', 'state', 'family_head']
    template_name = 'censo/persona/edit_person.html'
    success_url = reverse_lazy('personas')

    def form_valid(self, person_form):
        person = person_form.save(commit=False)

        # VALIDACIÓN 1: Solo puede haber un cabeza de familia por ficha
        if person.family_head:
            otros_jefes = Person.objects.filter(
                family_card=person.family_card,
                family_head=True,
                state=True
            ).exclude(pk=person.pk)

            if otros_jefes.exists():
                jefe_actual = otros_jefes.first()
                messages.error(
                    self.request,
                    f"Ya existe un cabeza de familia en esta ficha: {jefe_actual.full_name}. "
                    f"Primero debe cambiar el rol del cabeza actual."
                )
                return self.form_invalid(person_form)

        # VALIDACIÓN 2: El cabeza de familia debe ser mayor de 18 años
        if person.family_head and person.date_birth:
            from datetime import date
            today = date.today()
            age = today.year - person.date_birth.year - (
                (today.month, today.day) < (person.date_birth.month, person.date_birth.day)
            )

            if age < 18:
                messages.error(
                    self.request,
                    f"El cabeza de familia debe ser mayor de 18 años. "
                    f"La persona tiene {age} años."
                )
                return self.form_invalid(person_form)

        # VALIDACIÓN 3: El documento de identidad debe ser único
        if person.identification_person:
            duplicado = Person.objects.filter(
                identification_person=person.identification_person,
                state=True
            ).exclude(pk=person.pk).first()

            if duplicado:
                messages.error(
                    self.request,
                    f"El documento {person.identification_person} ya está registrado "
                    f"para {duplicado.full_name} en la ficha {duplicado.family_card.family_card_number}."
                )
                return self.form_invalid(person_form)

        # Guardar persona
        person_form.instance.user = self.request.user
        person.save()

        # LÓGICA: Desactivar ficha si no quedan miembros activos
        if not person.state:
            miembros_activos = Person.objects.filter(
                family_card_id=person.family_card_id,
                state=True
            ).count()

            if miembros_activos == 0:
                ficha = person.family_card
                ficha.state = False
                ficha.family_card_number = 0
                ficha.save()
                messages.warning(
                    self.request,
                    "La ficha familiar ha sido desactivada porque no quedan miembros activos."
                )

        messages.success(self.request, "Persona actualizada correctamente")
        return super(UpdatePerson, self).form_valid(person_form)

    def form_invalid(self, form):
        messages.error(
            self.request,
            "Por favor, corrija los errores en el formulario antes de continuar."
        )
        return super(UpdatePerson, self).form_invalid(form)


def person_by_gender(request):
    # Contar las personas por género
    cantidad = Person.objects.values('gender__gender')

    return JsonResponse({'cantidad': cantidad})


@require_POST
@csrf_protect
def update_family_head(request, family, person):
    # Aquí deberías validar permisos del usuario
    with transaction.atomic():
        family_obj = get_object_or_404(FamilyCard, id=family)
        person_obj = get_object_or_404(Person, id=person, family_card=family_obj)

        if person_obj.family_head:
            return JsonResponse(
                {'status': 'info', 'title': 'Información', 'message': 'La persona ya es el cabeza de familia.'})

        Person.objects.filter(family_card=family_obj, family_head=True).update(family_head=False)
        person_obj.family_head = True
        person_obj.save()
        return JsonResponse(
            {'status': 'success', 'title': 'Éxito', 'message': 'Cabeza de familia actualizado correctamente.'})


def delete_person_familyCard(request, person):
    # Aquí deberías validar permisos del usuario

    with transaction.atomic():
        old_person_obj = get_object_or_404(Person, id=person)
        old_family_card = old_person_obj.family_card

        nueva_ficha = FamilyCard.objects.create(
            address_home=old_family_card.address_home,
            sidewalk_home=old_family_card.sidewalk_home,
            latitude=old_family_card.latitude,
            longitude=old_family_card.longitude,
            zone=old_family_card.zone,
            organization=old_family_card.organization,
            family_card_number=FamilyCard.get_next_family_card_number(),
        )

        old_person_obj.family_card = nueva_ficha
        old_person_obj.family_head = True
        old_person_obj.save()

        return JsonResponse(
            {'status': 'success', 'title': 'Éxito', 'message': 'Persona desvinculada correctamente.'})



# Consultar los parámetros y retornar un json
@login_required
def get_system_parameters(request):
    """
    Vista API que retorna los parametros del sistema en formato JSON.
    Usa cache para mejorar el rendimiento.
    """
    from .utils import get_system_parameters_cached

    try:
        # Usar cache con timeout de 1 hora
        data = get_system_parameters_cached(timeout=3600)
        return JsonResponse(data)
    except Exception as e:
        logger.error(f"Error al obtener los parametros del sistema: {e}")
        return JsonResponse({'error': 'Error al obtener los parametros del sistema'}, status=500)


# Registro de Materiales de construcción
class MaterialConstructionView(LoginRequiredMixin, CreateView):
    model = MaterialConstructionFamilyCard
    form_class = MaterialConstructionFamilyForm
    template_name = 'censo/censo/material_construction_form.html'
    success_url = reverse_lazy('familyCardIndex')
    context_object_name = 'material_construction'
    extra_context = {'segment': 'material_construction'}

    def form_valid(self, form):
        # Guardado manual para asegurar que family_card se asigne correctamente
        pk = self.kwargs.get('pk')
        try:
            with transaction.atomic():
                instance = form.save(commit=False)
                # Asignar usuario si el modelo tiene ese campo (seguro no lo tiene en este modelo)
                try:
                    instance.user = self.request.user
                except Exception:
                    pass
                if pk:
                    family = get_object_or_404(FamilyCard, pk=pk)
                    instance.family_card = family
                instance.save()
                # guardar m2m si existieran
                form.save_m2m()
                # asignar la instancia creada a self.object para compatibilidad con get_success_url()
                self.object = instance
            messages.success(self.request, "Material de construcción creado correctamente")
            return redirect('familyCardIndex')
        except IntegrityError as e:
            # logger.error(f"Error al guardar MaterialConstructionFamilyCard: {e}")
            print(f"Error al guardar MaterialConstructionFamilyCard: {e}")
            form.add_error(None, "No se pudo guardar el registro por un conflicto en la base de datos.")
            messages.error(self.request, "Error al crear el material de construcción. Por favor, revise los campos.")
            return self.form_invalid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Error al crear el material de construcción. Por favor, revise los campos.")
        print(form.errors)
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        params = SystemParameters.objects.all().values('key', 'value')
        context['system_params'] = {p['key']: p['value'] for p in params}
        return context


# ============================================================================
# EXPORTACIÓN A EXCEL
# ============================================================================

@login_required
def export_persons_excel(request):
    """
    Exporta el listado de personas a Excel con el formato solicitado.
    Respeta el filtro de organización del usuario.

    Columnas:
    Nro de ficha, dirección, zona, tipo documento identidad, identificación,
    nombre 1, nombre 2, apellido 1, apellido 2, fecha nacimiento, eps,
    parentesco, género, estado civil, ocupación, nivel educativo, teléfono,
    cabeza de hogar
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from datetime import datetime

    try:
        # Query base de personas
        personas = Person.objects.select_related(
            'family_card',
            'family_card__sidewalk_home',
            'family_card__organization',
            'document_type',
            'gender',
            'eps',
            'kinship',
            'civil_state',
            'occupation',
            'education_level'
        ).filter(state=True)

        # Filtrar por organización del usuario
        if not (request.user.is_superuser or getattr(request, 'can_view_all', False)):
            user_organization = getattr(request, 'user_organization', None)
            if user_organization:
                personas = personas.filter(family_card__organization=user_organization)
            else:
                # Sin organización, retornar vacío
                messages.error(request, "No tiene una organización asignada.")
                return redirect('personas')

        # Ordenar por ficha y cabeza de familia
        personas = personas.order_by('family_card__family_card_number', '-family_head')

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Personas Registradas"

        # Estilos
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2196F3', end_color='2196F3', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        cell_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        # Headers según el orden solicitado
        headers = [
            'Nro de Ficha',
            'Dirección',
            'Zona',
            'Tipo Documento',
            'Identificación',
            'Nombre 1',
            'Nombre 2',
            'Apellido 1',
            'Apellido 2',
            'Fecha Nacimiento',
            'EPS',
            'Parentesco',
            'Género',
            'Estado Civil',
            'Ocupación',
            'Nivel Educativo',
            'Teléfono',
            'Cabeza de Hogar'
        ]

        # Escribir headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = cell_border

        # Ajustar anchos de columna
        column_widths = {
            'A': 15,  # Nro Ficha
            'B': 30,  # Dirección
            'C': 10,  # Zona
            'D': 15,  # Tipo Doc
            'E': 15,  # Identificación
            'F': 15,  # Nombre 1
            'G': 15,  # Nombre 2
            'H': 15,  # Apellido 1
            'I': 15,  # Apellido 2
            'J': 15,  # Fecha Nac
            'K': 25,  # EPS
            'L': 20,  # Parentesco
            'M': 12,  # Género
            'N': 15,  # Estado Civil
            'O': 20,  # Ocupación
            'P': 20,  # Nivel Educativo
            'Q': 15,  # Teléfono
            'R': 15   # Cabeza Hogar
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Escribir datos
        row_num = 2
        for persona in personas:
            # Determinar zona
            zona_text = ''
            if persona.family_card and persona.family_card.zone:
                zona_text = 'Urbana' if persona.family_card.zone == 'U' else 'Rural'

            # Cabeza de hogar
            cabeza_text = 'SÍ' if persona.family_head else 'NO'

            # Fecha de nacimiento formateada
            fecha_nac = persona.date_birth.strftime('%Y-%m-%d') if persona.date_birth else ''

            # Datos según el orden solicitado
            row_data = [
                persona.family_card.family_card_number if persona.family_card else '',
                persona.family_card.address_home if persona.family_card and persona.family_card.address_home else '',
                zona_text,
                persona.document_type.code_document_type if persona.document_type else '',
                persona.identification_person or '',
                persona.first_name_1 or '',
                persona.first_name_2 or '',
                persona.last_name_1 or '',
                persona.last_name_2 or '',
                fecha_nac,
                persona.eps.eps_name if persona.eps else '',
                persona.kinship.description_kinship if persona.kinship else '',
                persona.gender.gender if persona.gender else '',
                persona.civil_state.description_civil_state if persona.civil_state else '',
                persona.occupation.description_occupancy if persona.occupation else '',
                persona.education_level.description_education_level if persona.education_level else '',
                persona.cell_phone or '',
                cabeza_text
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = cell_border
                cell.alignment = Alignment(vertical='center')

            row_num += 1

        # Agregar filtros
        ws.auto_filter.ref = ws.dimensions

        # Congelar primera fila
        ws.freeze_panes = 'A2'

        # Crear response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # Nombre del archivo con fecha y organización
        org_name = ''
        if hasattr(request, 'user_organization') and request.user_organization:
            org_name = f"_{request.user_organization.organization_name.replace(' ', '_')}"

        filename = f'personas_registradas{org_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'

        wb.save(response)

        # Log de exportación
        print(f"Exportación Excel de personas: {personas.count()} registros - Usuario: {request.user.username}")

        return response

    except Exception as e:
        messages.error(request, f"Error al exportar a Excel: {str(e)}")
        return redirect('personas')

